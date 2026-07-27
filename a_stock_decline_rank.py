# -*- coding: utf-8 -*-
"""
A股全量股票涨跌幅排行榜
计算从指定日期至今的涨跌幅，按涨跌幅从小到大排序（涨最多/跌最少的排在前面）
输出包含股票名称、编码、市值、行业等信息

输出文件:
  - Markdown 报告: output/a_stock_decline_rank_<date>.md（可指定 --top 截断）
  - TXT 完整排名: output/a_stock_decline_rank_<date>.txt（始终全量，固定列宽）

数据源:
  - 股票列表: 巨潮 cninfo szse_stock.json (6213 只，单次 HTTP GET)
  - 实时行情/市值/PE: 腾讯财经 qt.gtimg.cn (批量查询，不封IP)
  - 历史K线(前复权): 腾讯财经 K线 API (多线程并发，不封IP)
  - 行业信息(可选): 东方财富 push2 clist (有风控，失败则标注"未知")

用法:
    python a_stock_decline_rank.py                          # 默认 2026-06-25
    python a_stock_decline_rank.py --date 2026-05-01        # 指定日期
    python a_stock_decline_rank.py --top 100                # MD只输出前100和后100
    python a_stock_decline_rank.py --workers 30             # 自定义并发数(默认20)
    python a_stock_decline_rank.py --no-cache               # 强制重新拉取
    python a_stock_decline_rank.py --no-industry            # 跳过行业查询(更快)
    python a_stock_decline_rank.py --no-txt                 # 不生成TXT完整排名
"""

import argparse
import json
import random
import sys
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

import requests

# ── 终端编码修复（Windows GBK → UTF-8） ───────────────────────────────
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ── 输出目录 ──────────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── 全局配置 ──────────────────────────────────────────────────────────
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36")

# ── 腾讯财经会话（不封IP，可高并发）───────────────────────────────────
TEN_SESSION = requests.Session()
TEN_SESSION.headers.update({"User-Agent": UA})
try:
    from requests.adapters import HTTPAdapter
    _ten_adapter = HTTPAdapter(pool_connections=40, pool_maxsize=80, max_retries=1)
    TEN_SESSION.mount("https://", _ten_adapter)
    TEN_SESSION.mount("http://", _ten_adapter)
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════════
#  第 1 步：获取全 A 股列表（巨潮 cninfo）
# ═══════════════════════════════════════════════════════════════════════

def get_stock_list_from_cninfo() -> list[dict]:
    """
    从巨潮 szse_stock.json 获取全 A 股股票列表（~6200 只）。
    返回每只股票: code, name, category
    """
    print("[1/4] 正在从巨潮拉取全 A 股列表...", flush=True)

    url = "http://www.cninfo.com.cn/new/data/szse_stock.json"
    req = urllib.request.Request(url)
    req.add_header("User-Agent", UA)
    try:
        resp = urllib.request.urlopen(req, timeout=15)
        data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"  [ERROR] 巨潮请求失败: {e}", flush=True)
        return []

    raw = data.get("stockList", [])
    stocks = []
    for s in raw:
        code = s.get("code", "")
        name = s.get("zwjc", "")  # 中文简称
        if not code or not name:
            continue
        # 过滤 B 股、退市股
        if "B" in s.get("category", ""):
            continue
        if "退" in name:
            continue
        stocks.append({
            "code": code,
            "name": name,
            "category": s.get("category", ""),
        })

    print(f"  ✓ 共获取 {len(stocks):,} 只 A 股\n", flush=True)
    return stocks


# ═══════════════════════════════════════════════════════════════════════
#  第 2 步：用腾讯实时行情批量获取市值/PE/当前价
# ═══════════════════════════════════════════════════════════════════════

def _tencent_prefix(code: str) -> str:
    """股票代码 → 腾讯行情前缀"""
    if code.startswith(("6", "9")):
        return "sh"
    elif code.startswith(("8", "4")):
        return "bj"
    else:
        return "sz"


def fill_market_data_tencent(stocks: list[dict]) -> int:
    """
    用腾讯 qt.gtimg.cn 批量获取实时行情（市值、PE、当前价）。
    腾讯单次 URL ~8KB，每批约 800 只。
    原地写入 cur_price, mcap, float_mcap, pe_ttm, pb, turnover_pct。
    返回成功数量。
    """
    print("[2/4] 正在用腾讯批量获取实时行情（市值/PE/当前价）...", flush=True)
    total = len(stocks)
    batch_size = 600  # URL 长度安全边界
    success = 0

    for batch_start in range(0, total, batch_size):
        batch_end = min(batch_start + batch_size, total)
        batch = stocks[batch_start:batch_end]

        # 构建批量查询 URL
        prefixed = [f"{_tencent_prefix(s['code'])}{s['code']}" for s in batch]
        url = "https://qt.gtimg.cn/q=" + ",".join(prefixed)

        try:
            req = urllib.request.Request(url)
            req.add_header("User-Agent", UA)
            resp = urllib.request.urlopen(req, timeout=15)
            raw = resp.read().decode("gbk", errors="replace")
        except Exception as e:
            print(f"  批次 {batch_start//batch_size+1} 请求失败: {e}", flush=True)
            continue

        # 解析每行返回值: v_sh600519="1~茅台~...~..."
        for line in raw.strip().split(";"):
            if "=" not in line or "~" not in line:
                continue
            try:
                key = line.split("=")[0].strip()
                code = key.split("_")[-1][2:]  # 提取纯6位代码
                vals = line.split('"')[1].split("~")
                if len(vals) < 53:
                    continue
            except Exception:
                continue

            # 找到对应股票
            # 用 code 匹配（batch 内线性查找，但 batch 只有 600 只，OK）
            for s in batch:
                if s["code"] == code:
                    s["cur_price"] = float(vals[3]) if vals[3] else 0.0
                    s["pe_ttm"] = float(vals[39]) if vals[39] else 0.0
                    s["mcap"] = float(vals[44]) * 1e8 if vals[44] else 0.0     # 亿→元
                    s["float_mcap"] = float(vals[45]) * 1e8 if vals[45] else 0.0
                    s["pb"] = float(vals[46]) if vals[46] else 0.0
                    s["turnover_pct"] = float(vals[38]) if vals[38] else 0.0
                    s["change_pct_today"] = float(vals[32]) if vals[32] else 0.0
                    success += 1
                    break

        done = batch_end
        print(f"  进度: {done:,}/{total:,} ({100*done/total:.1f}%)", flush=True)

    print(f"  ✓ 成功获取 {success:,} 只股票的实时行情\n", flush=True)
    return success


# ═══════════════════════════════════════════════════════════════════════
#  第 3 步：用腾讯 K线 API 获取参考日收盘价（多线程）
# ═══════════════════════════════════════════════════════════════════════

def _calc_before_metrics(date_close: dict, ref_date: str, ref_price: float,
                         cur_price: float) -> tuple[int, float, int]:
    """
    计算前导天数 & 前后天数比。

    逻辑: 从 ref_date 往前扫描，找到第一次达到与 ref_date→今日
    同等幅度（绝对值）变动的最早日期的交易天数，不论方向。
    即：如果 "之后" 跌了 -30%，找 "之前" 何时涨/跌过 ≥30%。

    返回: (before_days, ratio, after_days)
      before_days > 0  → 找到匹配的先前区间
      before_days = 0  → 涨跌幅 < 0.5%，忽略
      before_days = -N → 未找到（N = 可往前追溯的最大天数）
    """
    after_change = (cur_price / ref_price - 1) * 100
    target = abs(after_change)

    if target < 0.5:
        return 0, 0.0, 0

    all_dates = sorted(date_close.keys())
    if len(all_dates) < 3:
        return -1, 0.0, 0

    # after_days: trading days from ref_date to last available date
    after_days = sum(1 for d in all_dates if d > ref_date)
    if after_days <= 0:
        after_days = 1

    # 找到 ref_date 在排序列表中的位置
    ref_idx = None
    for i, d in enumerate(all_dates):
        if d >= ref_date:
            ref_idx = i
            break
    if ref_idx is None:
        ref_idx = len(all_dates) - 1

    # 往前扫描 —— 找绝对幅度 ≥ target 的最早日期（最近 ref_date 的）
    for i in range(ref_idx - 1, -1, -1):
        prior_price = date_close[all_dates[i]]
        if prior_price <= 0:
            continue
        prior_change_pct = abs((ref_price / prior_price - 1) * 100)

        if prior_change_pct >= target:
            before_days = ref_idx - i
            ratio = round(before_days / after_days, 2) if after_days > 0 else 0
            return before_days, ratio, after_days

    # 没找到：返回可追溯的最大天数（负数表示未找到）
    max_before = ref_idx
    if max_before <= 0:
        return -1, 0.0, after_days
    return -max_before, round(max_before / after_days, 2), after_days


def _find_ref_price(klines: list, ref_date: str):
    """
    在腾讯 K线数据中找 ref_date 的前复权收盘价。
    klines: [ [date, open, close, high, low, volume], ... ] 按时间升序
    返回: (close_price, actual_date) 或 (None, None)
    """
    if not klines:
        return None, None

    date_close = {}
    for bar in klines:
        if len(bar) < 3:
            continue
        d = str(bar[0])[:10]
        try:
            date_close[d] = float(bar[2])  # close
        except (ValueError, TypeError):
            continue

    if ref_date in date_close:
        return date_close[ref_date], ref_date

    # 往前找最近交易日
    try:
        parts = ref_date.split("-")
        dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None, None

    for i in range(1, 25):
        prev = (dt - timedelta(days=i)).strftime("%Y-%m-%d")
        if prev in date_close:
            return date_close[prev], prev

    return None, None


def _fetch_one_kline(code: str, ref_date: str) -> tuple[str, float | None, str | None, dict, list]:
    """
    获取参考日收盘价 + 全量 K线数据。
    返回: (code, ref_price, actual_date, date_close_dict, raw_klines)
    raw_klines: [ [date, open, close, high, low, volume], ... ] 原始K线列表
    """
    prefix = _tencent_prefix(code)
    url = (f"https://proxy.finance.qq.com/ifzqgtimg/appstock/app/fqkline/get"
           f"?param={prefix}{code},day,2020-01-01,,250,qfq")
    try:
        r = TEN_SESSION.get(
            url,
            headers={"Referer": "https://gu.qq.com/"},
            timeout=(5, 10),
        )
        if r.status_code != 200:
            return code, None, None, {}, []

        d = r.json()
        stock_data = d.get("data", {}).get(f"{prefix}{code}", {})
        klines = stock_data.get("qfqday") or stock_data.get("day") or []

        # 构建 date→close 映射
        date_close = {}
        for bar in klines:
            if len(bar) < 3:
                continue
            try:
                date_close[str(bar[0])[:10]] = float(bar[2])
            except (ValueError, TypeError):
                continue

        price, actual_date = _find_ref_price(klines, ref_date)
        return code, price, actual_date, date_close, klines
    except Exception:
        return code, None, None, {}, []


def fill_historical_prices(stocks: list[dict], ref_date: str, workers: int = 20) -> int:
    """
    用腾讯 K线 API（多线程）获取每只股票在 ref_date 的收盘价 + 计算前导指标。
    原地写入 ref_price / ref_date_actual / before_days / ratio / after_days。
    返回成功数量。
    """
    print(f"[3/4] 正在用腾讯K线获取历史数据 + 计算前导指标...", flush=True)
    total = len(stocks)
    success = 0
    t_start = time.time()

    batch_size = 200  # 每批提交200只

    for batch_start in range(0, total, batch_size):
        batch_end = min(batch_start + batch_size, total)
        batch = stocks[batch_start:batch_end]

        results: dict[str, tuple] = {}
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {
                pool.submit(_fetch_one_kline, s["code"], ref_date): s["code"]
                for s in batch
            }
            for future in as_completed(futures):
                try:
                    code, price, actual_date, date_close, raw_klines = future.result()
                    results[code] = (price, actual_date, date_close, raw_klines)
                except Exception:
                    pass

        # 写入结果 + 计算前导指标 + R/O评分
        for s in batch:
            res = results.get(s["code"])
            if res is None:
                continue
            price, actual_date, date_close, raw_klines = res
            if price is not None and price > 0:
                s["ref_price"] = price
                s["ref_date_actual"] = actual_date
                success += 1

                # 计算前导指标
                cur_p = s.get("cur_price", 0) or 0
                if cur_p > 0 and date_close:
                    before_days, ratio, after_days = _calc_before_metrics(
                        date_close, actual_date or ref_date, price, cur_p)
                    s["before_days"] = before_days
                    s["ratio"] = ratio
                    s["after_days"] = after_days

                # 计算 R/O 风险机会评分
                if cur_p > 0 and raw_klines:
                    try:
                        from risk_opportunity import compute_ro_scores
                        ro = compute_ro_scores(
                            raw_klines, actual_date or ref_date, cur_p)
                        s["ro_risk"] = ro["risk_score"]
                        s["ro_opp"] = ro["opportunity_score"]
                        s["ro_quadrant"] = ro["quadrant"]
                        s["ro_signals"] = ", ".join(ro.get("signals", [])[:3])
                    except Exception:
                        s["ro_risk"] = None
                        s["ro_opp"] = None
                        s["ro_quadrant"] = None
                        s["ro_signals"] = None

        # 进度
        done = batch_end
        elapsed = time.time() - t_start
        rate = done / elapsed if elapsed > 0 else 0
        eta = (total - done) / rate if rate > 0 else 0
        print(
            f"  进度: {done:,}/{total:,} ({100*done/total:.1f}%)  "
            f"成功={success:,}  速率={rate:.1f}只/秒  ETA={eta:.0f}s",
            flush=True,
        )

        if batch_end < total:
            time.sleep(0.2)

    elapsed = time.time() - t_start
    print(f"  ✓ 完成！成功={success:,}/{total:,}  耗时={elapsed:.1f}s\n", flush=True)
    return success


# ═══════════════════════════════════════════════════════════════════════
#  第 3.5 步(可选)：用东财补充行业信息
# ═══════════════════════════════════════════════════════════════════════

def fill_industry_eastmoney(stocks: list[dict]) -> int:
    """
    尝试从东财 push2 批量获取行业信息（沪深京分页拉取）。
    有风控风险，失败则行业字段保留"未知"。
    返回成功更新的数量。
    """
    print("[3.5/4] 正在从东财补充行业信息（可能因风控失败）...", flush=True)

    em_session = requests.Session()
    em_session.headers.update({
        "User-Agent": UA,
        "Referer": "https://quote.eastmoney.com/",
    })

    # 构建 code → stock 映射
    code_map = {s["code"]: s for s in stocks}
    updated = 0
    page = 1
    page_size = 2000

    while True:
        params = {
            "pn": str(page),
            "pz": str(page_size),
            "po": "1", "np": "1",
            "fltt": "2", "invt": "2",
            "fs": "m:0+t:6,m:0+t:13,m:0+t:80,m:0+t:81,m:1+t:2,m:1+t:23",
            "fields": "f12,f100",
        }
        try:
            r = em_session.get(
                "https://push2.eastmoney.com/api/qt/clist/get",
                params=params, timeout=15,
            )
            d = r.json()
        except Exception:
            print(f"  东财请求失败（风控/网络），跳过行业信息", flush=True)
            break

        data = d.get("data")
        if not data:
            break

        total = data.get("total", 0)
        items = data.get("diff") or []
        for it in items:
            code = it.get("f12", "")
            industry = (it.get("f100") or "").strip()
            if code in code_map and industry:
                code_map[code]["industry"] = industry
                updated += 1

        fetched = page * page_size
        print(f"  第 {page} 页: 累计 {min(fetched, total):,}/{total:,}  行业更新={updated:,}", flush=True)

        if fetched >= total:
            break
        page += 1
        time.sleep(0.6)  # 东财限流

    print(f"  ✓ 行业信息: {updated:,} 只已更新\n", flush=True)
    return updated


# ═══════════════════════════════════════════════════════════════════════
#  第 4 步：计算、排序、输出 Markdown
# ═══════════════════════════════════════════════════════════════════════

def _format_mcap(mcap_yuan: float) -> str:
    if mcap_yuan <= 0:
        return "-"
    yi = mcap_yuan / 1e8
    if yi >= 10000:
        return f"{yi / 10000:.2f}万亿"
    return f"{yi:.0f}亿"


def _format_price(p: float) -> str:
    if p <= 0:
        return "-"
    return f"{p:.2f}"


def generate_report(stocks: list[dict], ref_date: str, top_n: int = 0) -> str:
    """生成 Markdown 报告。"""
    print("[4/6] 正在生成MD报告...", flush=True)

    # 确保 industry 字段存在
    for s in stocks:
        s.setdefault("industry", "未知")
        s.setdefault("mcap", 0)
        s.setdefault("pe_ttm", 0)

    # 过滤有效数据
    valid = [s for s in stocks
             if (s.get("ref_price") or 0) > 0 and (s.get("cur_price") or 0) > 0]
    print(f"  有效股票数: {len(valid):,}（已剔除停牌/新股/数据缺失）", flush=True)

    # 计算涨跌幅
    for s in valid:
        s["change_pct"] = round((s["cur_price"] / s["ref_price"] - 1) * 100, 2)

    # 按涨跌幅从小到大排序（跌最少/涨最多 = 表现最好 → 排最前）
    valid.sort(key=lambda x: x["change_pct"])

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 统计 ──
    pcts = [s["change_pct"] for s in valid]
    up_count = sum(1 for p in pcts if p > 0)
    down_count = sum(1 for p in pcts if p < 0)
    sorted_pcts = sorted(pcts)
    median_pct = sorted_pcts[len(pcts) // 2] if pcts else 0

    lines = []
    lines.append("# A股全量股票涨跌幅排行榜")
    lines.append("")
    lines.append(f"| 项目 | 内容 |")
    lines.append(f"|------|------|")
    lines.append(f"| 生成时间 | {now_str} |")
    lines.append(f"| 基准日期 | {ref_date} |")
    lines.append(f"| 有效股票数 | {len(valid):,} 只 |")
    lines.append("")

    lines.append("## 市场概况")
    lines.append("")
    lines.append("| 指标 | 数值 |")
    lines.append("|------|------|")
    lines.append(f"| 上涨家数 | {up_count:,} ({100*up_count/len(valid):.1f}%) |")
    lines.append(f"| 下跌家数 | {down_count:,} ({100*down_count/len(valid):.1f}%) |")
    lines.append(f"| 中位数涨跌幅 | {median_pct:+.2f}% |")
    lines.append(f"| 平均涨跌幅 | {sum(pcts)/len(pcts):+.2f}% |")
    lines.append(f"| 最大涨幅 | {max(pcts):+.2f}% |")
    lines.append(f"| 最大跌幅 | {min(pcts):+.2f}% |")
    lines.append("")

    # 涨跌幅分布
    lines.append("### 涨跌幅分布")
    lines.append("")
    lines.append("| 区间 | 数量 | 占比 |")
    lines.append("|------|------|------|")
    for lo, hi in [(20, 9999), (10, 20), (5, 10), (2, 5), (0, 2),
                   (-2, 0), (-5, -2), (-10, -5), (-20, -10), (-9999, -20)]:
        cnt = sum(1 for p in pcts if lo <= p < hi)
        if hi >= 9999:
            label = f"≥ +{lo:.0f}%"
        elif lo <= -9999:
            label = f"≤ {hi:.0f}%"
        else:
            label = f"{lo:+.0f}% ~ {hi:+.0f}%"
        lines.append(f"| {label} | {cnt:,} | {100*cnt/len(valid):.1f}% |")
    lines.append("")

    # ── 排名表 ──
    show_full = (top_n <= 0 or top_n * 2 >= len(valid))

    if show_full:
        lines.append(f"## 完整排名（共 {len(valid):,} 只，按涨跌幅从小到大）")
        lines.append("")
        lines.append(_table_header())
        for i, s in enumerate(valid, 1):
            lines.append(_table_row(i, s))
    else:
        lines.append(f"## TOP {top_n} — 涨幅最大 / 最抗跌（表现最好）")
        lines.append("")
        lines.append(_table_header())
        for i, s in enumerate(reversed(valid[-top_n:]), 1):
            lines.append(_table_row(i, s))

        lines.append("")
        lines.append(f"## BOTTOM {top_n} — 跌幅最大（表现最差）")
        lines.append("")
        lines.append(_table_header())
        for i, s in enumerate(valid[:top_n], 1):
            lines.append(_table_row(i, s))

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("*数据来源: 巨潮(股票列表) + 腾讯财经(行情/市值/PE/历史K线) + 东方财富(行业)*")
    lines.append("")
    lines.append("*免责声明: 本报告仅供研究参考，不构成投资建议。股市有风险，投资需谨慎。*")

    return "\n".join(lines)


def _table_header() -> str:
    return (
        "| 排名 | 代码 | 名称 | 行业 | 总市值 | 基准日价 | 当前价 | 涨跌幅 | 前导天 | 天数比 | 风险 | 机会 | 象限 | PE |"
        "\n|------|------|------|------|------|---------|--------|--------|--------|--------|------|------|------|----|"
    )


def _fmt_before_days(d: int | None) -> str:
    """格式化前导天数"""
    if d is None:
        return "-"
    if d == 0:
        return "-"
    if d < 0:
        return f">{abs(d)}"
    return str(d)


def _table_row(rank: int, s: dict) -> str:
    pct = s["change_pct"]
    if pct > 5:
        emoji = "🟢"
    elif pct > 0:
        emoji = "🟡"
    elif pct > -5:
        emoji = "🟠"
    else:
        emoji = "🔴"

    pe = "-"
    pe_val = s.get("pe_ttm", 0)
    if pe_val and pe_val > 0:
        pe = f"{pe_val:.1f}"
    elif pe_val and pe_val < 0:
        pe = "亏损"

    before = _fmt_before_days(s.get("before_days"))
    ratio_val = s.get("ratio")
    if ratio_val is not None and ratio_val > 0:
        ratio_str = f"{ratio_val:.1f}x"
    else:
        ratio_str = "-"

    # R/O 列
    ro_risk = s.get("ro_risk")
    ro_opp = s.get("ro_opp")
    ro_quad = s.get("ro_quadrant", "")
    risk_str = f"{ro_risk:.0f}" if ro_risk is not None else "-"
    opp_str = f"{ro_opp:.0f}" if ro_opp is not None else "-"
    quad_map = {"green": "🟢", "yellow": "🟡", "white": "⚪", "red": "🔴"}
    quad_str = quad_map.get(ro_quad, "-") if ro_quad else "-"

    return (
        f"| {rank} | {s['code']} | {s['name']} | {s.get('industry','')} | "
        f"{_format_mcap(s.get('mcap', 0))} | {_format_price(s.get('ref_price', 0))} | "
        f"{_format_price(s.get('cur_price', 0))} | {emoji} {pct:+.2f}% | "
        f"{before} | {ratio_str} | {risk_str} | {opp_str} | {quad_str} | {pe} |"
    )


# ═══════════════════════════════════════════════════════════════════════
#  TXT 完整排名输出（固定列宽，全量）
# ═══════════════════════════════════════════════════════════════════════

def _pad(s: str, width: int, align: str = "<") -> str:
    """按宽度填充字符串（中文字符占2个宽度）"""
    # 计算实际显示宽度：中文占2，ASCII占1
    disp = 0
    for ch in str(s):
        disp += 2 if '一' <= ch <= '鿿' or '　' <= ch <= '〿' or '＀' <= ch <= '￯' else 1
    pad_total = width - disp
    if pad_total <= 0:
        return str(s)
    if align == "<":
        return str(s) + " " * pad_total
    elif align == ">":
        return " " * pad_total + str(s)
    else:  # "^"
        left = pad_total // 2
        return " " * left + str(s) + " " * (pad_total - left)


def generate_txt_report(stocks: list[dict], ref_date: str) -> str:
    """
    生成完整的 TXT 排名文件（含所有有效股票，不受 --top 限制）。
    固定列宽，每列对齐，方便在文本编辑器/终端中查看。
    """
    # 过滤有效数据 + 计算涨跌幅
    valid = [s for s in stocks
             if (s.get("ref_price") or 0) > 0 and (s.get("cur_price") or 0) > 0]
    for s in valid:
        s["change_pct"] = round((s["cur_price"] / s["ref_price"] - 1) * 100, 2)

    # 按涨跌幅从小到大排序
    valid.sort(key=lambda x: x["change_pct"])

    # 统计
    pcts = [s["change_pct"] for s in valid]
    up_count = sum(1 for p in pcts if p > 0)
    down_count = sum(1 for p in pcts if p < 0)
    sorted_pcts = sorted(pcts)
    median_pct = sorted_pcts[len(pcts) // 2]

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 列宽定义 ──
    COL_RANK = 6
    COL_CODE = 8
    COL_NAME = 12
    COL_INDUSTRY = 14
    COL_MCAP = 10
    COL_REF = 10
    COL_CUR = 10
    COL_PCT = 10
    COL_BEFORE = 8
    COL_RATIO = 8
    COL_RISK = 6
    COL_OPP = 6
    COL_QUAD = 5
    COL_PE = 8

    def fmt_mcap(v):
        if v <= 0: return "-"
        yi = v / 1e8
        return f"{yi:.0f}亿" if yi < 10000 else f"{yi/10000:.2f}万亿"

    def fmt_price(v):
        return f"{v:.2f}" if v > 0 else "-"

    def fmt_before(d):
        if d is None: return "-"
        if d == 0: return "-"
        if d < 0: return f">{abs(d)}"
        return str(d)

    def fmt_ratio(r):
        if r is None or r == 0: return "-"
        return f"{r:.1f}x"

    def fmt_pe(v):
        if v is None or v == 0: return "-"
        if v < 0: return "亏损"
        return f"{v:.1f}"

    lines = []
    # ── 表头 ──
    sep = "+" + "+".join(["-" * (w + 1) for w in [
        COL_RANK, COL_CODE, COL_NAME, COL_INDUSTRY, COL_MCAP,
        COL_REF, COL_CUR, COL_PCT, COL_BEFORE, COL_RATIO,
        COL_RISK, COL_OPP, COL_QUAD, COL_PE]]) + "+"

    lines.append(sep)
    header = "| " + " | ".join([
        _pad("排名", COL_RANK, "^"),
        _pad("代码", COL_CODE, "^"),
        _pad("名称", COL_NAME, "^"),
        _pad("行业", COL_INDUSTRY, "^"),
        _pad("总市值", COL_MCAP, "^"),
        _pad("基准日价", COL_REF, "^"),
        _pad("当前价", COL_CUR, "^"),
        _pad("涨跌幅", COL_PCT, "^"),
        _pad("前导天", COL_BEFORE, "^"),
        _pad("天数比", COL_RATIO, "^"),
        _pad("风险", COL_RISK, "^"),
        _pad("机会", COL_OPP, "^"),
        _pad("象限", COL_QUAD, "^"),
        _pad("PE", COL_PE, "^"),
    ]) + " |"
    lines.append(header)
    lines.append(sep.replace("+", "|").replace("-", "="))

    def fmt_ro(v):
        return f"{v:.0f}" if v is not None else "-"

    def fmt_quad(q):
        return {"green": "G", "yellow": "Y", "white": "W", "red": "R"}.get(q, "-") if q else "-"

    # ── 数据行 ──
    for i, s in enumerate(valid, 1):
        row = "| " + " | ".join([
            _pad(str(i), COL_RANK, ">"),
            _pad(s["code"], COL_CODE, "^"),
            _pad(s["name"], COL_NAME, "<"),
            _pad(s.get("industry", "未知"), COL_INDUSTRY, "<"),
            _pad(fmt_mcap(s.get("mcap", 0)), COL_MCAP, ">"),
            _pad(fmt_price(s.get("ref_price", 0)), COL_REF, ">"),
            _pad(fmt_price(s.get("cur_price", 0)), COL_CUR, ">"),
            _pad(f"{s['change_pct']:+.2f}%", COL_PCT, ">"),
            _pad(fmt_before(s.get("before_days")), COL_BEFORE, ">"),
            _pad(fmt_ratio(s.get("ratio")), COL_RATIO, ">"),
            _pad(fmt_ro(s.get("ro_risk")), COL_RISK, ">"),
            _pad(fmt_ro(s.get("ro_opp")), COL_OPP, ">"),
            _pad(fmt_quad(s.get("ro_quadrant")), COL_QUAD, "^"),
            _pad(fmt_pe(s.get("pe_ttm")), COL_PE, ">"),
        ]) + " |"
        lines.append(row)

    lines.append(sep)

    # ── 前导信息 ──
    header_lines = [
        f"A股全量股票涨跌幅排行榜",
        f"基准日期: {ref_date}  截止日期: {datetime.now().strftime('%Y-%m-%d')}  生成时间: {now_str}",
        f"有效股票数: {len(valid):,}  上涨: {up_count:,}  下跌: {down_count:,}  中位数: {median_pct:+.2f}%  平均: {sum(pcts)/len(pcts):+.2f}%",
        f"最大涨幅: {max(pcts):+.2f}%  最大跌幅: {min(pcts):+.2f}%",
        "",
    ]

    return "\n".join(header_lines + lines)


# ═══════════════════════════════════════════════════════════════════════
#  Excel 完整排名输出
# ═══════════════════════════════════════════════════════════════════════

def generate_excel_report(stocks: list[dict], ref_date: str, xlsx_path: Path):
    """
    生成完整的 Excel 排名文件（所有有效股票，带格式化）。
    使用 pandas + openpyxl，含冻结表头、自动列宽、条件着色。
    """
    import pandas as pd

    # 过滤有效数据 + 计算涨跌幅
    valid = [s for s in stocks
             if (s.get("ref_price") or 0) > 0 and (s.get("cur_price") or 0) > 0]
    for s in valid:
        s["change_pct"] = round((s["cur_price"] / s["ref_price"] - 1) * 100, 2)
    valid.sort(key=lambda x: x["change_pct"])

    # 构建 DataFrame
    rows = []
    for i, s in enumerate(valid, 1):
        pe = s.get("pe_ttm", 0) or 0
        bd = s.get("before_days")
        rt = s.get("ratio", 0) or 0
        ro_q = s.get("ro_quadrant", "")
        rows.append({
            "排名": i,
            "代码": s["code"],
            "名称": s["name"],
            "行业": s.get("industry", "未知"),
            "总市值(亿)": round(s.get("mcap", 0) / 1e8, 2) if s.get("mcap") else 0,
            "基准日收盘价": s.get("ref_price", 0),
            "当前价": s.get("cur_price", 0),
            "涨跌幅(%)": s["change_pct"],
            "前导天数": bd if bd and bd > 0 else (abs(bd) if bd and bd < 0 else 0),
            "前导标记": ">" + str(abs(bd)) if bd and bd < 0 else ("-" if bd == 0 else ""),
            "天数比": rt,
            "风险分": s.get("ro_risk"),
            "机会分": s.get("ro_opp"),
            "R/O象限": {"green": "🟢高机会低风险", "yellow": "🟡高机会高风险",
                      "white": "⚪低机会低风险", "red": "🔴低机会高风险"}.get(ro_q, ""),
            "R/O信号": s.get("ro_signals", ""),
            "PE(TTM)": pe if pe > 0 else None,
        })

    df = pd.DataFrame(rows)

    # 写入 Excel
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="涨跌幅排名", index=False, startrow=4)

        ws = writer.sheets["涨跌幅排名"]

        # ── 标题行 ──
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=1, column=1, value=f"A股全量股票涨跌幅排行榜")
        ws.cell(row=2, column=1,
                value=f"基准日期: {ref_date}  截止日期: {datetime.now().strftime('%Y-%m-%d')}  "
                      f"生成时间: {now_str}  有效股票数: {len(valid):,}")

        # ── 冻结表头 ──
        ws.freeze_panes = "A6"

        DATA_START = 6       # 数据从第6行开始（第5行=表头）
        HEADER_ROW = 5       # 表头在第5行

        # ── 自动列宽 ──
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    text = str(cell.value)
                    length = sum(2 if ord(ch) > 127 else 1 for ch in text)
                    max_len = max(max_len, length)
            ws.column_dimensions[ws.cell(row=HEADER_ROW, column=col_idx).column_letter].width = min(max_len + 3, 30)

        # ── 涨跌幅列条件着色 ──
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        pct_col = 8  # 涨跌幅列 (1-indexed)
        for row_idx in range(DATA_START, len(valid) + DATA_START):
            cell = ws.cell(row=row_idx, column=pct_col)
            if cell.value is not None:
                val = float(cell.value)
                if val > 5:
                    cell.fill = green_fill
                elif val > 0:
                    cell.fill = yellow_fill
                elif val < -5:
                    cell.fill = red_fill
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center")

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 数据区域边框
        for row_idx in range(DATA_START, len(valid) + DATA_START):
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).border = thin_border


# ═══════════════════════════════════════════════════════════════════════
#  缓存机制
# ═══════════════════════════════════════════════════════════════════════

def _cache_path(ref_date: str) -> Path:
    return OUTPUT_DIR / f"stock_data_{ref_date}.json"


def load_cache(ref_date: str) -> list[dict] | None:
    path = _cache_path(ref_date)
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list) and len(data) > 500:
            return data
    except Exception:
        pass
    return None


def save_cache(ref_date: str, stocks: list[dict]):
    path = _cache_path(ref_date)
    slim = []
    keys = ["code", "name", "category", "cur_price", "pe_ttm", "mcap",
            "float_mcap", "pb", "industry", "ref_price", "ref_date_actual",
            "before_days", "ratio", "after_days",
            "ro_risk", "ro_opp", "ro_quadrant", "ro_signals"]
    for s in stocks:
        slim.append({k: s.get(k) for k in keys})
    path.write_text(json.dumps(slim, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✓ 缓存已保存: {path}\n", flush=True)


# ═══════════════════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="A股全量股票涨跌幅排行榜 — 从指定日期计算至今的涨跌幅",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python a_stock_decline_rank.py
  python a_stock_decline_rank.py --date 2026-06-25
  python a_stock_decline_rank.py --date 2026-06-01 --top 100
  python a_stock_decline_rank.py --workers 30 --top 50
  python a_stock_decline_rank.py --no-cache --no-industry   # 最快模式
        """,
    )
    parser.add_argument("--date", "-d", default="2026-06-25",
                        help="基准日期 YYYY-MM-DD（默认: 2026-06-25）")
    parser.add_argument("--top", "-t", type=int, default=0,
                        help="只输出前N和后N名（0=全部，默认0）")
    parser.add_argument("--workers", "-w", type=int, default=20,
                        help="腾讯K线并发线程数（默认20，最大50）")
    parser.add_argument("--no-cache", action="store_true",
                        help="忽略缓存，强制重新拉取")
    parser.add_argument("--no-industry", action="store_true",
                        help='跳过东财行业查询（更快，行业标为"未知"）')
    parser.add_argument("--no-txt", action="store_true",
                        help="不生成TXT完整排名文件（默认生成）")
    parser.add_argument("--no-excel", action="store_true",
                        help="不生成Excel排名文件（默认生成）")
    parser.add_argument("--output", "-o", default=None,
                        help="MD输出文件路径（默认: output/a_stock_decline_rank_<date>.md）")
    args = parser.parse_args()

    ref_date = args.date
    top_n = args.top
    workers = min(args.workers, 50)
    use_cache = not args.no_cache
    fetch_industry = not args.no_industry

    print("=" * 60)
    print(f"  A股全量股票涨跌幅排行榜")
    print(f"  基准日期: {ref_date}")
    print(f"  截止日期: {datetime.now().strftime('%Y-%m-%d')}")
    print(f"  并发线程: {workers}")
    print(f"  行业查询: {'是' if fetch_industry else '否（跳过东财）'}")
    print("=" * 60)
    print()

    # ── 加载缓存 ──
    stocks = None
    if use_cache:
        stocks = load_cache(ref_date)
        if stocks:
            print(f"[CACHE] 加载缓存: {len(stocks):,} 只股票\n", flush=True)

    # ── Step 1: 股票列表（巨潮） ──
    if stocks is None:
        stocks = get_stock_list_from_cninfo()
        if not stocks:
            print("[ERROR] 无法获取股票列表")
            sys.exit(1)

    # ── Step 2: 实时行情（腾讯批量） ──
    has_market = any(s.get("cur_price", 0) > 0 for s in stocks)
    if not has_market:
        n = fill_market_data_tencent(stocks)
        if n == 0:
            print("[ERROR] 无法获取实时行情")
            sys.exit(1)

    # ── Step 3: 历史K线（腾讯多线程） ──
    has_historical = any(s.get("ref_price") is not None for s in stocks)
    if not has_historical:
        n = fill_historical_prices(stocks, ref_date, workers=workers)
        if n == 0:
            print("[ERROR] 无法获取历史价格")
            sys.exit(1)

    # ── Step 3.5: 行业信息（东财，可选） ──
    has_industry = any(s.get("industry", "") not in ("", "未知", None) for s in stocks)
    if fetch_industry and not has_industry:
        fill_industry_eastmoney(stocks)

    # 确保 industry 默认值
    for s in stocks:
        if not s.get("industry"):
            s["industry"] = "未知"

    # ── 保存缓存 ──
    save_cache(ref_date, stocks)

    # ── Step 4: 生成报告 ──
    report = generate_report(stocks, ref_date, top_n)

    # ── 写入文件 ──
    out_path = Path(args.output or (OUTPUT_DIR / f"a_stock_decline_rank_{ref_date}.md"))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(report, encoding="utf-8")
    print(f"  ✓ MD报告已保存至: {out_path}")
    size_kb = out_path.stat().st_size / 1024
    print(f"  ✓ 文件大小: {size_kb:.0f} KB")
    print()

    # ── Step 5: 生成TXT完整排名（始终全量） ──
    if not args.no_txt:
        print("[5/6] 正在生成TXT完整排名...", flush=True)
        txt_report = generate_txt_report(stocks, ref_date)
        txt_path = out_path.with_suffix(".txt")
        txt_path.write_text(txt_report, encoding="utf-8")
        print(f"  ✓ TXT完整排名已保存至: {txt_path}")
        print(f"  ✓ 文件大小: {txt_path.stat().st_size / 1024:.0f} KB")
        print()

    # ── Step 6: 生成Excel完整排名（始终全量） ──
    if not args.no_excel:
        print("[6/6] 正在生成Excel完整排名...", flush=True)
        xlsx_path = out_path.with_suffix(".xlsx")
        generate_excel_report(stocks, ref_date, xlsx_path)
        print(f"  ✓ Excel完整排名已保存至: {xlsx_path}")
        print(f"  ✓ 文件大小: {xlsx_path.stat().st_size / 1024:.0f} KB")
        print()

    # ── 终端摘要 ──
    valid = [s for s in stocks
             if s.get("ref_price") and s.get("cur_price", 0) > 0]
    if valid:
        valid.sort(key=lambda x: x["cur_price"] / x["ref_price"] - 1)
        print("─" * 60)
        print("  涨幅 TOP 10（表现最好）:")
        for i, s in enumerate(reversed(valid[-10:]), 1):
            pct = (s["cur_price"] / s["ref_price"] - 1) * 100
            print(f"    {i:2d}. {s['code']} {s['name']:<10s} {pct:+.2f}%  |  {s.get('industry','')}")
        print("─" * 60)
        print("  跌幅 TOP 10（表现最差）:")
        for i, s in enumerate(valid[:10], 1):
            pct = (s["cur_price"] / s["ref_price"] - 1) * 100
            print(f"    {i:2d}. {s['code']} {s['name']:<10s} {pct:+.2f}%  |  {s.get('industry','')}")
        print("─" * 60)

    print("\nDone!")


if __name__ == "__main__":
    main()
